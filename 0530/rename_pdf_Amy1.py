import os
import re
import glob
import numpy as np

try:
    import easyocr
    import pdfplumber
    import openpyxl  # 新增 Excel 讀取套件
except ImportError:
    print("請先安裝必備套件: pip install easyocr pdfplumber openpyxl")
    exit()

# 初始化 EasyOCR 引擎
print("正在初始化 EasyOCR 辨識引擎，請稍候...")
reader = easyocr.Reader(['ch_tra', 'en'], gpu=False)

def load_vendor_mapping_from_excel():
    """
    從同目錄下的 Excel (.xlsx) 檔案匯入廠商全名(B欄)與簡稱(C欄)對照表
    """
    excel_file = "vendor_mapping.xlsx"
    vendor_dict = {}
    
    if not os.path.exists(excel_file):
        print(f"❌ 錯誤：在同目錄下找不到 【{excel_file}】 檔案，請確認檔名是否正確！")
        return vendor_dict

    try:
        # 開啟 Excel 檔案
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active # 預設讀取第一個工作表
        
        # 從第二行開始讀取 (跳過第一行的標題「名稱」「簡稱」)
        for row in range(2, sheet.max_row + 1):
            full_name_cell = sheet.cell(row=row, column=2).value  # B 欄 (Column 2) -> 名稱
            short_name_cell = sheet.cell(row=row, column=3).value # C 欄 (Column 3) -> 簡稱
            
            if full_name_cell and short_name_cell:
                full_name = str(full_name_cell).strip()
                short_name = str(short_name_cell).strip()
                
                if full_name and short_name:
                    vendor_dict[full_name] = short_name
                    
        print(f"📊 成功從 Excel 匯入 {len(vendor_dict)} 筆廠商簡稱對照資料。")
    except Exception as e:
        print(f"❌ 讀取 Excel 對照表出錯 (請確認檔案是否被 Excel 開啟中並鎖定): {e}")
        
    return vendor_dict

def extract_po_info_from_pdf(pdf_path, vendor_dict):
    """
    智慧更名核心：結合 Excel 簡稱資料庫與雙區特徵判定。
    """
    po_number = None
    vendor_name = ""
    is_supplement = False
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return None
            
            first_page = pdf.pages[0]
            width = first_page.width
            height = first_page.height
            
            # 區塊一：左上半部資訊區 -> 專抓單號與廠商
            crop_main_bbox = (0, 0, width * 0.8, height * 0.55)
            main_page = first_page.crop(crop_main_bbox)
            img_main = main_page.to_image(resolution=300)
            img_main_np = np.array(img_main.original)
            
            main_results = reader.readtext(img_main_np, detail=0)
            main_text = " ".join(main_results)
            
            # 區塊二：右上角手寫區
            crop_top_right = (width * 0.65, height * 0.12, width, height * 0.45)
            tr_page = first_page.crop(crop_top_right)
            img_tr = tr_page.to_image(resolution=300) 
            img_tr_np = np.array(img_tr.original)
            tr_text = "".join(reader.readtext(img_tr_np, detail=0)).replace(" ", "")
            
            # 區塊三：最底部手寫備註區
            crop_bottom = (width * 0.5, height * 0.85, width, height)
            bt_page = first_page.crop(crop_bottom)
            img_bt = bt_page.to_image(resolution=300)
            img_bt_np = np.array(img_bt.original)
            bt_text = "".join(reader.readtext(img_bt_np, detail=0)).replace(" ", "")
            
            filename = os.path.basename(pdf_path)
            tr_text_cleaned = re.sub(r'(單據日期|公司傳真|\d{4}/\d{2}/\d{2}|\d{2}-\d{4}-\d{4}|頁次)', '', tr_text)
            
            # 【1】提取採購單號
            po_match = re.search(r'\b48\d{8}\b', main_text)
            if po_match:
                po_number = po_match.group(0).strip()
            else:
                backup_po = re.search(r'\b\d{10}\b', main_text)
                if backup_po:
                    po_number = backup_po.group(0).strip()

            # 【2】動態比對 Excel 匯入的 B欄全名，匹配成功就換成 C欄簡稱
            matched_full_name = None
            clean_main_text = main_text.replace(" ", "")
            
            for full_name in vendor_dict.keys():
                if full_name in clean_main_text:
                    matched_full_name = full_name
                    break
                    
            if matched_full_name:
                vendor_name = vendor_dict[matched_full_name]
            else:
                # 備援：若 Excel 沒有匹配到，抓取「廠商名稱:」後面的前 4 個字當暫時名稱
                vendor_match = re.search(r'廠商名稱\s*[:：\s]*(.*?)(?:\s|$)', main_text)
                if vendor_match:
                    vendor_name = re.sub(r'[:：\s\-]', "", vendor_match.group(1).strip())[:4]
                else:
                    vendor_name = "未知廠商"

            # 【3】精密補單判定
            combined_target_text = tr_text_cleaned + bt_text
            supplement_keywords = ['補', '單', '布', '車', '革', '草', '畢', '軍', '目', '田', '早', '痛']
            match_count = sum(1 for char in combined_target_text if char in supplement_keywords)
            
            if "補" in combined_target_text or match_count >= 2:
                is_supplement = True
            else:
                is_supplement = False
                
            return po_number, vendor_name, is_supplement
                
    except Exception as e:
        print(f"處理文件 {os.path.basename(pdf_path)} 時出錯: {e}")
    
    return None

def rename_pdfs_to_po(folder_path):
    if not os.path.exists(folder_path):
        print(f"錯誤：找不到指定的資料夾路徑 -> {folder_path}")
        return

    # 讀取 Excel 對照表
    vendor_dict = load_vendor_mapping_from_excel()
    if not vendor_dict:
        print("⚠️ 警告：廠商對照表為空，將影響簡稱轉換功能。")

    pdf_files = glob.glob(os.path.join(folder_path, "*.pdf"))
    if not pdf_files:
        print("未在指定路徑找到任何 PDF 文件。")
        return

    print(f"\n目標資料夾: {folder_path}")
    print(f"開始執行【Excel 動態資料庫更名】...\n")
    success_count = 0
    
    for pdf_path in pdf_files:
        old_name = os.path.basename(pdf_path)
        info = extract_po_info_from_pdf(pdf_path, vendor_dict)
        
        if info and info[0]:
            po_num, v_name, is_supp = info
            
            if is_supp:
                new_name = f"{v_name}-{po_num}-補.pdf"
            else:
                new_name = f"{v_name}-{po_num}.pdf"
                
            new_path = os.path.join(folder_path, new_name)
            
            counter = 1
            while os.path.exists(new_path):
                if pdf_path == new_path:
                    break
                if is_supp:
                    new_name = f"{v_name}-{po_num}-補_{counter}.pdf"
                else:
                    new_name = f"{v_name}-{po_num}_{counter}.pdf"
                new_path = os.path.join(folder_path, new_name)
                counter += 1
                
            if pdf_path == new_path:
                continue
                
            try:
                os.rename(pdf_path, new_path)
                print(f"[V] 更名成功: {old_name} -> {new_name}")
                success_count += 1
            except Exception as e:
                print(f"[X] 重命名失敗 {old_name}: {e}")
        else:
            print(f"[?] 無法從 {old_name} 辨識出有效採購單號，保持原狀。")

    print(f"\n處理完成！本次成功自動更名/校正了 {success_count} / {len(pdf_files)} 個檔案。")

if __name__ == "__main__":
    # 您的 PDF 採購單存放路徑
    target_directory = r"D:\shuyu.lin\Documents\採購單PDF"
    rename_pdfs_to_po(target_directory)